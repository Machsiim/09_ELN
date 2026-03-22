from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from ..schemas import GenerateSampleRequest
from ..utils import extract_template_headers

router = APIRouter(tags=["Template"])


@router.post("/generate-sample-excel")
async def generate_sample_excel(request: GenerateSampleRequest):
    """Generate an empty Excel file with column headers from a template schema."""
    headers = extract_template_headers(request.template_schema)
    if not headers:
        raise HTTPException(status_code=400, detail="Keine Felder im Template-Schema gefunden.")

    wb = Workbook()
    ws = wb.active
    ws.title = request.template_name[:31]  # Excel sheet name max 31 chars

    # Style definitions
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Write headers
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        # Auto-width based on header length
        ws.column_dimensions[cell.column_letter].width = max(15, len(header) + 4)

    # Freeze header row
    ws.freeze_panes = "A2"

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    safe_name = request.template_name.replace('"', "").replace("/", "_").replace("\\", "_")
    filename = f"Vorlage_{safe_name}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
