from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from typing import List, Dict, Any, Optional, Tuple
from io import BytesIO, StringIO
import csv

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

router = APIRouter(tags=["Export"])


# Section color palette (dark / mid / light) — mirrors the frontend's
# $section-colors map in measurement-series-detail.scss so the exported
# Excel uses the same per-section colors as the on-screen filter / table view.
SECTION_COLOR_PALETTE: List[Dict[str, str]] = [
    {"dark": "00649C", "mid": "B3D4E8", "light": "E3EFF6"},  # 0 blue
    {"dark": "6D8C17", "mid": "C8DEA0", "light": "E9F1D6"},  # 1 green
    {"dark": "8B5E00", "mid": "E0C88A", "light": "F3EAD2"},  # 2 gold
    {"dark": "7B2D5F", "mid": "D4A3C4", "light": "F0DFE9"},  # 3 magenta
    {"dark": "2D6E6E", "mid": "A3CECE", "light": "DCEEED"},  # 4 teal
    {"dark": "8B3A3A", "mid": "D4A3A3", "light": "F0DFDF"},  # 5 red
    {"dark": "3A4A6E", "mid": "AAB5CC", "light": "DDE2EC"},  # 6 slate
    {"dark": "4A2D7B", "mid": "BFA3D4", "light": "E6DCF0"},  # 7 deep purple
    {"dark": "A83A7A", "mid": "E0A3C9", "light": "F5DFEC"},  # 8 pink
    {"dark": "1A1A1A", "mid": "9A9A9A", "light": "E0E0E0"},  # 9 gray
]


class ExportRequest(BaseModel):
    data: List[Dict[str, Any]] = Field(..., description="Zeilen als Liste von Dictionaries")
    columns: List[str] = Field(..., description="Spaltenreihenfolge")
    column_sections: Optional[Dict[str, str]] = Field(default=None, description="Zuordnung Spaltenname -> Sektionsname")
    column_cards: Optional[Dict[str, str]] = Field(default=None, description="Zuordnung Spaltenname -> Kartenname")
    column_field_labels: Optional[Dict[str, str]] = Field(default=None, description="Zuordnung Spaltenname -> reines Feld-Label")
    column_section_colors: Optional[Dict[str, int]] = Field(
        default=None,
        description="Zuordnung Spaltenname -> ColorIndex (0..9) der Section. Wenn nicht gesetzt, wird die Default-Blau-Palette fuer alle Spalten verwendet.",
    )
    meta_columns: Optional[List[str]] = Field(default=None, description="Spalten, die ueber alle Header-Zeilen gespannt werden (Mess-ID, Erstellt von, ...)")
    filename: str = Field(default="Export", description="Dateiname ohne Erweiterung")
    sheet_name: str = Field(default="Daten", description="Sheet-Name fuer Excel")


def _build_ranges(values: List[str]) -> List[Tuple[str, int, int]]:
    """Return list of (value, start_col, end_col) where consecutive equal entries are merged.

    Empty strings always stay as 1-column ranges (no merging across empty cells).
    """
    ranges: List[Tuple[str, int, int]] = []
    if not values:
        return ranges

    current = values[0]
    start = 1
    for idx, val in enumerate(values[1:], start=2):
        if val != current or val == "":
            ranges.append((current, start, idx - 1))
            current = val
            start = idx
    ranges.append((current, start, len(values)))
    return ranges


@router.post("/export/excel")
async def export_excel(request: ExportRequest):
    """Generate a formatted Excel file from data rows.

    Header layout (matches frontend table):
      row 1: section (merged across all its columns)
      row 2: card    (merged across all its columns within a section)
      row 3: field label
    Meta columns (e.g. Mess-ID, Erstellt von, Erstellt am) span all 3 header rows.
    """
    if not request.columns:
        raise HTTPException(status_code=400, detail="Keine Spalten angegeben.")

    wb = Workbook()
    ws = wb.active
    ws.title = request.sheet_name[:31]

    # Default (fallback) palette — used when no column_section_colors is provided
    DEFAULT_DARK = "2F5496"
    DEFAULT_MID = "3A6BB0"
    DEFAULT_LIGHT = "4472C4"

    white_font = Font(bold=True, color="FFFFFF", size=11)
    meta_fill = PatternFill(start_color=DEFAULT_DARK, end_color=DEFAULT_DARK, fill_type="solid")
    header_fallback_fill = PatternFill(start_color=DEFAULT_LIGHT, end_color=DEFAULT_LIGHT, fill_type="solid")
    zebra_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

    column_color_indices = request.column_section_colors or {}

    def _palette_for_column(col: str) -> Dict[str, str]:
        idx = column_color_indices.get(col)
        if idx is None:
            return {"dark": DEFAULT_DARK, "mid": DEFAULT_MID, "light": DEFAULT_LIGHT}
        return SECTION_COLOR_PALETTE[idx % len(SECTION_COLOR_PALETTE)]

    def _solid(hex_color: str) -> PatternFill:
        return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

    def _dark_font(hex_color: str) -> Font:
        return Font(bold=True, color=hex_color, size=11)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def _data_border(is_section_boundary: bool) -> Border:
        return Border(
            left=Side(style="thin"),
            right=Side(style="medium" if is_section_boundary else "thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

    sections = request.column_sections or {}
    cards = request.column_cards or {}
    field_labels = request.column_field_labels or {}
    meta_set = set(request.meta_columns or [])

    has_sections = len(sections) > 0
    data_start_row = 4 if has_sections else 2
    section_boundary_cols: set = set()

    if has_sections:
        # Build per-column section / card values.
        # For meta columns we set empty values so they don't get merged with neighbours.
        section_values = [
            "" if col in meta_set else sections.get(col, "")
            for col in request.columns
        ]
        card_values = [
            "" if col in meta_set else cards.get(col, "")
            for col in request.columns
        ]

        # Section boundary columns: right edge of every named section that has
        # another named section to its right. Used to draw a thicker vertical
        # separator between sections in the header and data rows.
        section_ranges_all = _build_ranges(section_values)
        for i, (value, _start_col, end_col) in enumerate(section_ranges_all):
            if not value:
                continue
            if any(r[0] for r in section_ranges_all[i + 1:]):
                section_boundary_cols.add(end_col)

        # Row 1: sections (merged). Each section gets its own dark palette color.
        for value, start_col, end_col in _build_ranges(section_values):
            if not value:
                continue
            # All columns in this range share the same section, hence same palette
            palette = _palette_for_column(request.columns[start_col - 1])
            section_fill = _solid(palette["dark"])
            if start_col != end_col:
                ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
            cell = ws.cell(row=1, column=start_col, value=value)
            cell.font = white_font
            cell.fill = section_fill
            cell.alignment = center
            for c in range(start_col, end_col + 1):
                ws.cell(row=1, column=c).border = thin_border
                ws.cell(row=1, column=c).fill = section_fill

        # Row 2: cards (merged within section). Cards use the section's mid color.
        for value, start_col, end_col in _build_ranges(card_values):
            if not value:
                continue
            palette = _palette_for_column(request.columns[start_col - 1])
            card_fill = _solid(palette["mid"])
            if start_col != end_col:
                ws.merge_cells(start_row=2, start_column=start_col, end_row=2, end_column=end_col)
            cell = ws.cell(row=2, column=start_col, value=value)
            cell.font = _dark_font(palette["dark"])
            cell.fill = card_fill
            cell.alignment = center
            for c in range(start_col, end_col + 1):
                ws.cell(row=2, column=c).border = thin_border
                ws.cell(row=2, column=c).fill = card_fill

        # Row 3: field labels. Fields use the section's light color.
        for col_idx, col_name in enumerate(request.columns, start=1):
            if col_name in meta_set:
                continue
            palette = _palette_for_column(col_name)
            label = field_labels.get(col_name, col_name)
            cell = ws.cell(row=3, column=col_idx, value=label)
            cell.font = _dark_font(palette["dark"])
            cell.fill = _solid(palette["light"])
            cell.alignment = center
            cell.border = thin_border

        # Meta columns: merge vertically across all 3 header rows (always dark blue)
        for col_idx, col_name in enumerate(request.columns, start=1):
            if col_name not in meta_set:
                continue
            ws.merge_cells(start_row=1, start_column=col_idx, end_row=3, end_column=col_idx)
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = white_font
            cell.fill = meta_fill
            cell.alignment = center
            for r in range(1, 4):
                ws.cell(row=r, column=col_idx).border = thin_border
                ws.cell(row=r, column=col_idx).fill = meta_fill
    else:
        # Fallback: single header row, no section/card info
        for col_idx, col_name in enumerate(request.columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = white_font
            cell.fill = header_fallback_fill
            cell.alignment = center
            cell.border = thin_border

    # Apply medium-weight right border on section boundary columns across header rows
    for col_idx in section_boundary_cols:
        for r in range(1, data_start_row):
            ws.cell(row=r, column=col_idx).border = Border(
                left=Side(style="thin"),
                right=Side(style="medium"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

    # Data rows (with zebra striping and section separators)
    for row_idx, row in enumerate(request.data, start=data_start_row):
        is_zebra = (row_idx - data_start_row) % 2 == 1
        for col_idx, col_name in enumerate(request.columns, start=1):
            value = row.get(col_name)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = _data_border(col_idx in section_boundary_cols)
            if is_zebra:
                cell.fill = zebra_fill

    # Auto-width: base on the field-label / column-name, not on section/card titles
    for col_idx, col_name in enumerate(request.columns, start=1):
        label = field_labels.get(col_name, col_name) if has_sections else col_name
        max_len = len(str(label))
        for row in request.data:
            val = row.get(col_name)
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)

    ws.freeze_panes = f"A{data_start_row}"

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    safe_name = request.filename.replace('"', "").replace("/", "_").replace("\\", "_")

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{safe_name}.xlsx"'},
    )


@router.post("/export/csv")
async def export_csv(request: ExportRequest):
    """Generate a CSV file from data rows.

    Header layout matches the Excel/frontend view: section row, card row, field-label row.
    For meta columns the column name is repeated on the field-label row and left blank on
    the section/card rows.
    """
    if not request.columns:
        raise HTTPException(status_code=400, detail="Keine Spalten angegeben.")

    output = StringIO()
    writer = csv.writer(output)

    sections = request.column_sections or {}
    cards = request.column_cards or {}
    field_labels = request.column_field_labels or {}
    meta_set = set(request.meta_columns or [])

    has_sections = len(sections) > 0

    if has_sections:
        section_row = [
            "" if col in meta_set else sections.get(col, "")
            for col in request.columns
        ]
        card_row = [
            "" if col in meta_set else cards.get(col, "")
            for col in request.columns
        ]
        field_row = [
            col if col in meta_set else field_labels.get(col, col)
            for col in request.columns
        ]
        writer.writerow(section_row)
        writer.writerow(card_row)
        writer.writerow(field_row)
    else:
        writer.writerow(request.columns)

    for row in request.data:
        writer.writerow([row.get(col, "") for col in request.columns])

    csv_bytes = output.getvalue().encode("utf-8-sig")
    buffer = BytesIO(csv_bytes)

    safe_name = request.filename.replace('"', "").replace("/", "_").replace("\\", "_")

    return StreamingResponse(
        buffer,
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{safe_name}.csv"'},
    )
