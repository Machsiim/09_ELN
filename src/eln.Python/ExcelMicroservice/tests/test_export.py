from fastapi.testclient import TestClient
from app.main import app
from openpyxl import load_workbook
import io

client = TestClient(app)


def _payload():
    return {
        "data": [
            {"Mess-ID": 1, "Erstellt von": "alice", "Erstellt am": "01.05.2026",
             "Probe - Masse": 12.3, "Probe - Volumen": 4.5, "Notiz": "ok"},
            {"Mess-ID": 2, "Erstellt von": "bob",   "Erstellt am": "02.05.2026",
             "Probe - Masse": 10.0, "Probe - Volumen": 3.1, "Notiz": "n/a"},
        ],
        "columns": ["Mess-ID", "Erstellt von", "Erstellt am",
                    "Probe - Masse", "Probe - Volumen", "Notiz"],
        "column_sections": {
            "Mess-ID": "Allgemein", "Erstellt von": "Allgemein", "Erstellt am": "Allgemein",
            "Probe - Masse": "Messung", "Probe - Volumen": "Messung", "Notiz": "Messung",
        },
        "column_cards": {
            "Mess-ID": "Mess-ID", "Erstellt von": "Erstellt von", "Erstellt am": "Erstellt am",
            "Probe - Masse": "Probe", "Probe - Volumen": "Probe", "Notiz": "Notiz",
        },
        "column_field_labels": {
            "Mess-ID": "Mess-ID", "Erstellt von": "Erstellt von", "Erstellt am": "Erstellt am",
            "Probe - Masse": "Masse", "Probe - Volumen": "Volumen", "Notiz": "Notiz",
        },
        "meta_columns": ["Mess-ID", "Erstellt von", "Erstellt am"],
        "filename": "test", "sheet_name": "Test",
    }


def test_export_excel_three_row_header():
    r = client.post("/export/excel", json=_payload())
    assert r.status_code == 200

    wb = load_workbook(io.BytesIO(r.content))
    ws = wb.active

    # Field labels in row 3 — meta cols are vertically merged so the value lives in row 1
    row3 = [ws.cell(row=3, column=c).value for c in range(1, 7)]
    assert row3 == [None, None, None, "Masse", "Volumen", "Notiz"]

    # Section in row 1 col 1 = Mess-ID (meta merged vertically), col 4 = Messung
    assert ws.cell(row=1, column=1).value == "Mess-ID"
    assert ws.cell(row=1, column=4).value == "Messung"

    # Card in row 2 col 4 = Probe (covers cols 4-5)
    assert ws.cell(row=2, column=4).value == "Probe"

    # Data starts at row 4
    assert ws.cell(row=4, column=1).value == 1
    assert ws.cell(row=4, column=4).value == 12.3

    # Meta column is merged vertically across rows 1..3 in column 1
    merged_ranges = {str(r) for r in ws.merged_cells.ranges}
    assert "A1:A3" in merged_ranges
    # "Probe" card merged across cols 4-5 in row 2
    assert "D2:E2" in merged_ranges
    # "Messung" section merged across cols 4-6 in row 1
    assert "D1:F1" in merged_ranges


def test_export_csv_three_header_rows():
    r = client.post("/export/csv", json=_payload())
    assert r.status_code == 200

    text = r.content.decode("utf-8-sig")
    lines = text.splitlines()

    # 3 header rows + 2 data rows
    assert len(lines) == 5
    # Section row: meta cells empty
    assert lines[0].startswith(",,,Messung,Messung,Messung")
    # Card row
    assert lines[1] == ",,,Probe,Probe,Notiz"
    # Field row
    assert lines[2] == "Mess-ID,Erstellt von,Erstellt am,Masse,Volumen,Notiz"
